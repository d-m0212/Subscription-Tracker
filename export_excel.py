from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from database import get_all_subscriptions, get_metrics, get_upcoming_renewals
from datetime import datetime

def generate_excel_report():
    """Generate Excel report with charts and formatted data"""
    
    # Get data
    subscriptions = get_all_subscriptions()
    metrics = get_metrics()
    renewals = get_upcoming_renewals()
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Summary Sheet
    ws_summary = wb.create_sheet('Summary')
    create_summary_sheet(ws_summary, metrics, subscriptions)
    
    # Create Subscriptions Sheet
    ws_subs = wb.create_sheet('All Subscriptions')
    create_subscriptions_sheet(ws_subs, subscriptions)
    
    # Create Renewals Sheet
    ws_renewals = wb.create_sheet('Upcoming Renewals')
    create_renewals_sheet(ws_renewals, renewals)
    
    # Save file
    filename = f'subscription_insights_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    
    return filename

def create_summary_sheet(ws, metrics, subscriptions):
    """Create summary sheet with metrics and charts"""
    
    # Title
    ws['A1'] = 'Subscription Spending Insights'
    ws['A1'].font = Font(size=18, bold=True)
    ws.merge_cells('A1:D1')
    
    # Date
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(size=10, italic=True)
    
    # Key Metrics
    ws['A4'] = 'Key Metrics'
    ws['A4'].font = Font(size=14, bold=True)
    
    ws['A5'] = 'Total Monthly Cost:'
    ws['B5'] = f"₹{metrics['total_monthly']}"
    ws['B5'].font = Font(bold=True, color='0000FF')
    
    ws['A6'] = 'Total Annual Cost:'
    ws['B6'] = f"₹{metrics['total_annual']}"
    ws['B6'].font = Font(bold=True, color='008000')
    
    ws['A7'] = 'Active Subscriptions:'
    ws['B7'] = metrics['total_subscriptions']
    ws['B7'].font = Font(bold=True)
    
    # Category Breakdown
    ws['A9'] = 'Spending by Category'
    ws['A9'].font = Font(size=14, bold=True)
    
    ws['A10'] = 'Category'
    ws['B10'] = 'Monthly Cost'
    ws['C10'] = 'Percentage'
    
    # Header styling
    for col in ['A10', 'B10', 'C10']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    row = 11
    for category, amount in metrics['categories'].items():
        percentage = (amount / metrics['total_monthly'] * 100) if metrics['total_monthly'] > 0 else 0
        ws[f'A{row}'] = category
        ws[f'B{row}'] = amount
        ws[f'C{row}'] = f"{percentage:.1f}%"
        row += 1
    
    # Pie Chart - Category Breakdown
    pie = PieChart()
    pie.title = "Spending by Category"
    pie.height = 10
    pie.width = 15
    
    data = Reference(ws, min_col=2, min_row=10, max_row=row-1)
    labels = Reference(ws, min_col=1, min_row=11, max_row=row-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    
    ws.add_chart(pie, 'E9')
    
    # Bar Chart - Top Subscriptions
    ws['A' + str(row + 2)] = 'Top Subscriptions by Monthly Cost'
    ws['A' + str(row + 2)].font = Font(size=14, bold=True)
    
    # Sort subscriptions by monthly cost
    sorted_subs = sorted(subscriptions, key=lambda x: x['monthly_cost'], reverse=True)[:10]
    
    ws['A' + str(row + 3)] = 'Subscription'
    ws['B' + str(row + 3)] = 'Monthly Cost'
    
    for col in ['A' + str(row + 3), 'B' + str(row + 3)]:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    chart_row = row + 4
    for sub in sorted_subs:
        ws[f'A{chart_row}'] = sub['name']
        ws[f'B{chart_row}'] = sub['monthly_cost']
        chart_row += 1
    
    # Bar Chart
    bar = BarChart()
    bar.title = "Top Subscriptions by Monthly Cost"
    bar.height = 12
    bar.width = 20
    
    data = Reference(ws, min_col=2, min_row=row+3, max_row=chart_row-1)
    labels = Reference(ws, min_col=1, min_row=row+4, max_row=chart_row-1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    
    ws.add_chart(bar, 'E' + str(row + 10))
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def create_subscriptions_sheet(ws, subscriptions):
    """Create detailed subscriptions sheet"""
    
    # Title
    ws['A1'] = 'All Subscriptions'
    ws['A1'].font = Font(size=16, bold=True)
    
    # Headers
    headers = ['Name', 'Amount', 'Billing Cycle', 'Category', 'Start Date', 'Renewal Date', 'Monthly Cost']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, sub in enumerate(subscriptions, start=4):
        ws.cell(row=row, column=1, value=sub['name'])
        ws.cell(row=row, column=2, value=sub['amount'])
        ws.cell(row=row, column=3, value=sub['billing_cycle'].capitalize())
        ws.cell(row=row, column=4, value=sub['category'])
        ws.cell(row=row, column=5, value=sub['start_date'])
        ws.cell(row=row, column=6, value=sub['renewal_date'])
        ws.cell(row=row, column=7, value=round(sub['monthly_cost'], 2))
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

def create_renewals_sheet(ws, renewals):
    """Create upcoming renewals sheet"""
    
    # Title
    ws['A1'] = 'Upcoming Renewals (Next 90 Days)'
    ws['A1'].font = Font(size=16, bold=True)
    
    if not renewals:
        ws['A3'] = 'No upcoming renewals in the next 90 days'
        return
    
    # Headers
    headers = ['Name', 'Amount', 'Billing Cycle', 'Category', 'Renewal Date', 'Days Until Renewal']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data with conditional formatting
    for row, renewal in enumerate(renewals, start=4):
        ws.cell(row=row, column=1, value=renewal['name'])
        ws.cell(row=row, column=2, value=renewal['amount'])
        ws.cell(row=row, column=3, value=renewal['billing_cycle'].capitalize())
        ws.cell(row=row, column=4, value=renewal['category'])
        ws.cell(row=row, column=5, value=renewal['renewal_date'])
        
        days_cell = ws.cell(row=row, column=6, value=renewal['days_until'])
        
        # Highlight urgent renewals (within 7 days)
        if renewal['days_until'] <= 7:
            days_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            days_cell.font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
